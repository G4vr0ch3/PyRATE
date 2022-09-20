#!/usr/bin/python3.10


################################################################################
#                                                                              #
#                                                                              #
#                   GNU AFFERO GENERAL PUBLIC LICENSE                          #
#                       Version 3, 19 November 2007                            #
#                                                                              #
#    This programms aims at sanitizing .xlsx, .xlsm and .xls documents.        #
#    Copyright (C) 2022 Gavroche, Roxane                                       #
#                                                                              #
#    This program is free software: you can redistribute it and/or modify      #
#    it under the terms of the GNU Affero General Public License as published  #
#    by the Free Software Foundation, either version 3 of the License, or      #
#    (at your option) any later version.                                       #
#                                                                              #
#     This program is distributed in the hope that it will be useful,          #
#    but WITHOUT ANY WARRANTY; without even the implied warranty of            #
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the             #
#    GNU Affero General Public License for more details.                       #
#                                                                              #
#    You should have received a copy of the GNU Affero General Public License  #
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.    #
#                                                                              #
#                                                                              #
################################################################################


import xlsxwriter

from openpyxl import Workbook


################################################################################


# Fetching data from source
def get_sheets(path):

    # Storing document data
    workbook = Workbook(path)

    # Retrieve sheet names
    sheet_names=workbook.sheetnames

    sheets = []

    # Retrieve each sheet content
    for sheet_index in range (len(sheet_names)):
        val = []

        sheet = workbook.sheet_by_index(sheet_index)

        # Determine sheet dimensions
        r = sheet.nrows
        c = sheet.ncols

        dim = (r,c)

        # Fetch content cell by cell
        for rw in range(r):
            for cl in range(c):
                val.append(sheet.cell_value(rw, cl))

        sheets.append([sheet_names[sheet_index], dim, val])

    return sheets


################################################################################


# Creating sanitized document
def recompose(workbook, sheets):

    # Creating each sheet
    for sheet_index in range(len(sheets)):

        worksheet = workbook.add_worksheet(sheets[sheet_index][0])

        cell = 0

        # Filling cells one by one based on sheet dimensions
        for rw in range((sheets[sheet_index][1][0])):
            for cl in range((sheets[sheet_index][1][1])):

                worksheet.write(rw, cl, sheets[sheet_index][2][cell])

                cell += 1


################################################################################


def sanitiz(path):

    info('Sanitizing ' + path.split("/")[-1])

    # Fetching data from source
    try:
        sheets = get_sheets(path)

    except:
        fail('Data extraction failed')
        return False, ''

    info('Creating new document')

    # Creating sanitized document
    try:
        workbook = xlsxwriter.Workbook('Outputs/out_' + path.split("/")[-1].split(".")[0] + ".xlsx")

        recompose(workbook, sheets)

        workbook.close()

    except:
        fail('Document recomposition failed')
        return False, ''

    success('Document sanitized successfully.')
    return True, 'Outputs/out_' + path.split("/")[-1].split(".")[0] + ".xlsx"


################################################################################


if __name__ == "__main__":
    print('Please run main.py or read software documentation')

    exit()

else:
    from .prints import *
